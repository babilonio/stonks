import requests
import pandas as pd
from datetime import datetime
from selenium import webdriver
import time
import os
import shutil
from openpyxl import load_workbook




# Copy here the stockrow search result url 
url = "https://stockrow.com/screener/7ac6d4e6-b723-4905-8493-5502ef577361"






def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name, 
            startrow=startrow if startrow is not None else 0, 
            **to_excel_kwargs)
        return
    
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)
    
    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)
    
    # copy existing sheets
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

driver = webdriver.Chrome()
driver.get(url)
time.sleep(2) # Let the user actually see something!
element = driver.find_element_by_xpath("//*[@id=\"root\"]/div/div/section/div/div[2]/section/div[2]/div/div/div/div[2]/div/div/div[3]/div/div")

companies = element.text.split('\n')[1:]

shutil.rmtree(url.split('/')[-1], ignore_errors=True)
os.mkdir(url.split('/')[-1])
os.chdir(url.split('/')[-1])

try:
    os.remove("database.xlsx") 
except OSError:
    pass

driver.close()
counter=0

for company in companies:

    print("\n ------------   DOWNLOADING ", company, " DATA ------------\n")

    base_url = "https://stockrow.com/api/companies/{company}/financials.xlsx?dimension=A&section={type}&sort=desc"
    types = ["Income Statement", "Balance Sheet", "Cash Flow"]

    for t in types:
        url = base_url.format(company=company, type=t)
        r = requests.get(url, allow_redirects=True)
        open(company + "_" + t + ".xlsx", 'wb').write(r.content)

    df = pd.DataFrame()

    for t in types:
        df = df.append(pd.read_excel(company + "_" + t + ".xlsx", engine="openpyxl"))
        os.remove(company + "_" + t + ".xlsx") 
    if (len(df.columns[1:]) < 10):
        print("\t SKIPPING! Company data not old enough...")

        continue
    
    counter +=1
    # ****START PRICES SECTION
    print("\t fetchin prices from Yahoo...")

    timestamp = int(datetime.timestamp(datetime.now()))
    download_url = "https://query1.finance.yahoo.com/v7/finance/download/{company}?period1=1325203200&period2={time}&interval=1mo&events=history&includeAdjustedClose=true".format(company=company.replace('.','-'), time=timestamp)

    r = requests.get(download_url, allow_redirects=True)
    open('historical_data.csv', 'wb').write(r.content)
    historical_data = pd.read_csv('historical_data.csv')
    os.remove("historical_data.csv") 

    latest = historical_data.tail(1)['Open']

    historical_data = historical_data[pd.to_datetime(historical_data['Date']).dt.month == 1]
    historical_data = historical_data.transpose()

    print("\t appending to stockrow data...")

    prices = [p for p in reversed(historical_data.loc['Open'].values)]    
    _df = pd.DataFrame([prices + [0]*(10 - len(prices))], columns = df.columns[1:])
    _df["Unnamed: 0"] = 'Stock Price'
    df = df.append(_df)

    _df2 = pd.DataFrame([list(latest.values)*10], columns = df.columns[1:])
    _df2["Unnamed: 0"] = 'Current Price'
    df = df.append(_df2)
    # **** END PRICES SECTION

    print("\t reordering...")

    df[company] = company
    df = df[[df.columns[-1], *df.columns[:-1]]]

    df['field_id'] = df.apply(lambda row: row[company] + " | " + row["Unnamed: 0"], axis=1)
    df = df[[*df.columns[:2], df.columns[-1], *reversed(df.columns[2:-1])]]

    df = df.rename(columns={"Unnamed: 0": 'field_name'}, errors="raise")
    df["Counter"] = counter
    df = df[[df.columns[-1], *df.columns[:-1]]]

    df = df.drop_duplicates()

    append_df_to_excel("database.xlsx", df, index=False)



